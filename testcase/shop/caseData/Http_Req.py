import openpyxl
import json
# from common.os_path import case_dir # 引入测试用例的路径
# from common.request import Http_Request
import random
import unittest
from utils.log import log
from interfaces.wxshop.MallAction import MallAction
from testcase.shop.sql.appletWX import wx_applet_evaluate
from testcase.shop.caseData.applet_data import submit_order
from testcase.shop.sql.webMP import mp_label
from utils import runlevel, timestamp
from ddt import data, unpack, ddt
from faker import Faker
import unittest
import inspect
import json

# class Case:
#
#     def __init__(self):  # 初始化测试数据 访问接口只需要用到：url data  header，初始值为None，读取EXCEL中对应的值，再赋上
#         self.case_id = None
#         self.url = None
#         self.data = None
#         self.header = None
#         self.title = None
#         self.method = None
#         self.expected = None
#         self.actual = None
#         self.result = None
#
#
# class DoExcel:
#
#
#     def __init__(self,file_name):
#         try:
#             self.workbook = openpyxl.load_workbook(filename=file_name) # 使用python中openpyxl打开excel
#         except FileNotFoundError as e:
#             print("Excel不存在:{}".format(e))
#             raise e
#
#
#     def get_case(self,sheet_name):
#         sheet = self.workbook[sheet_name] # 根据打开的excel 定位excel中的表单sheet
#         max_row = sheet.max_row # 获取sheet表单中的最大行
#
#         test_data = [] # 定义一个空的list装载用例
#         for i in range(2,max_row+1):
#             datas =Case() # 实例化上面的Case，下面要读取excel中的测试数据，赋值给初始None
#             datas.case_id = sheet.cell(row = i ,column = 1).value # 读取excel中case_id
#             datas.url = sheet.cell(row = i ,column = 2).value # 读取excel中url
#             datas.data = sheet.cell(row = i ,column = 3).value # 读取excel中data
#             datas.header = sheet.cell(row = i ,column = 4).value # 读取excel中header
#             datas.title = sheet.cell(row = i ,column = 5).value # 读取excel中title
#             datas.method = sheet.cell(row = i ,column = 6).value # 读取excel中请求方式
#             datas.expected = sheet.cell(row = i ,column = 7).value # 读取excel中期望结果
#             test_data.append(datas) # 读取出来的测试数据放到list
#
#         return test_data



import requests

class Http_Request:
    def __init__(self,method,url,data,header=None,cookie=None):
        if method == 'post':
            self.resp = requests.post(url=url,data=data,headers=header,cookies=cookie)
        elif method == 'get':
            self.resp = requests.get(url=url,data=data,headers=header,cookies=cookie)

        elif method == 'put':
            self.resp = requests.put(url=url,data=data,headers=header,cookies=cookie)

    def get_status_code(self):
        return self.resp.status_code
    def get_json(self):
        return self.resp.json()

    def get_text(self):
        return self.resp.text



if __name__ == '__main__':

    class add_evaluate(unittest.TestCase):
        def setUp(self) -> None:
            """
            测试前数据准备
            :return:
            """
            self.api = MallAction()
            self.api.set_user(mobile=15198034727)
            self.db = wx_applet_evaluate()
            self.faker = Faker('zh_CN')

    url = 'http://dev-gateway.worldfarm.com/wx-mall/web/order/list'
    data = {
            "pn":1,
            "ps":2,
            "orderStatus":10
            }

    header2 = {"Accept":"application/json",
               "Accept-Encoding":"gzip,deflate",
               "User-Agent":"Mozilla/5.0 (Macintosh; Intel Mac OS X 10_13_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/68.0.3440.106 Safari/537.36",
               "_App-Version_":"2.3.1",
               "_Device-Id_":"cc4feebe419791332bbcff5e0fdf084a",
               "_Device-Type_":"iOS",
               "_Token_":"eyJhbGciOiJIUzUxMiJ9.eyJqdGkiOiJjYzRmZWViZTQxOTc5MTMzMmJiY2ZmNWUwZmRmMDg0YSIsImRldmljZUdyb3VwIjoiVVNFUiIsImlwIjoiMTkyLjE2OC44OS4xNzUiLCJpc3MiOiJaWVBfU1NPIiwidXNlciI6IntcImFjY291bnRUeXBlXCI6MjEsXCJhcHBJZFwiOlwiMTFcIixcImRldmljZVR5cGVcIjpcIkFORFJPSUQtVVNFUlwiLFwiaGVhZEltZ1wiOlwiaHR0cDovL3p5cC1mYXJtLTIub3NzLWFwLXNvdXRoZWFzdC0xLmFsaXl1bmNzLmNvbS9kYXRhL2ZjLXVzZXIvaGVhZEltZy8xNTkzNTgyNTM0ODAzLmpwZ1wiLFwiaWRcIjo2OTQsXCJuZXdcIjpmYWxzZSxcInBhc3N3b3JkXCI6XCI3Mzc1MzA0MWY3NzZiNjg0MjE0NmFlNTA0MTQ4NjBiMzdmMzY5Nzg2Njk4N2NiMGZcIixcInBob25lXCI6XCIxNTM4ODEyNjA3MlwiLFwic3RhdHVzXCI6MSxcInVzZXJOYW1lXCI6XCLlvKDkuInlkbXlkbXlkbVcIn0iLCJpYXQiOjE2MDcwNjUzNDV9.-Uy_6KyXi-3QNqyLCwnjylczcDIYQPa3NfMOsI-cPUD0bLp_uSDzrmy6Auz5fLuBZ-YtqeklwACGgf26D9fUlA",
               "Accept_Language":"zh",
               "region":"online"}

    t = Http_Request('post',url=url,data=data,header=header2)
    print(t.get_text())










